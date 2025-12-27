import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

# Estilos comunes
header_font = Font(bold=True, color='FFFFFF')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_sheet(wb, name, color, headers, data, is_first=False):
    if is_first:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)

    header_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    return ws

headers = ['Método', 'Endpoint', 'Descripción', 'Auth', 'Rate Limit', 'Request Body', 'Response', 'Llamado por']

# ============ API GATEWAY (Puerto 8080) ============
gateway_data = [
    ['GET', '/health', 'Health check', 'No', 'No', 'N/A', '{"status": "ok"}', 'Externo'],
    ['POST', '/auth/register', 'Registrar usuario', 'No', '10/min', '{"phone", "nombre", "apellidos"}', '{"user_id", "message"}', 'App móvil'],
    ['POST', '/auth/send-code', 'Enviar código SMS', 'No', '10/min', '{"phone"}', '{"code", "expires_in"}', 'App móvil'],
    ['POST', '/auth/verify', 'Verificar código -> tokens', 'No', '10/min', '{"phone", "code", "device_id", "device_type"}', '{"access_token", "refresh_token"}', 'App móvil'],
    ['GET', '/api/v1/me', 'Datos usuario actual', 'JWT', '100/min', 'N/A', '{"id", "phone", "nombre", "apellidos"}', 'App móvil'],
    ['GET', '/api/v1/me/sessions', 'Sesiones activas', 'JWT', '100/min', 'N/A', '[{"id", "device_type", "created_at"}]', 'App móvil'],
    ['GET', '/api/v1/me/stats', 'Estadísticas usuario', 'JWT', '100/min', 'N/A', '{"total_messages", "threats_detected"}', 'App móvil'],
    ['POST', '/api/v1/me/logout', 'Cerrar sesión', 'JWT', '100/min', 'N/A', '{"message"}', 'App móvil'],
    ['POST', '/api/v1/me/logout-all', 'Cerrar todas sesiones', 'JWT', '100/min', 'N/A', '{"message"}', 'App móvil'],
    ['GET', '/api/v1/conversations', 'Listar conversaciones', 'JWT', '100/min', 'N/A', '[{"id", "title", "message_count"}]', 'App móvil'],
    ['POST', '/api/v1/conversations', 'Crear conversación', 'JWT', '100/min', '{"title"}', '{"id", "title", "created_at"}', 'App móvil'],
    ['GET', '/api/v1/conversations/{id}', 'Obtener conversación', 'JWT', '100/min', 'N/A', '{"id", "title", "messages"}', 'App móvil'],
    ['GET', '/api/v1/conversations/{id}/messages', 'Mensajes conversación', 'JWT', '100/min', 'N/A', '[{"role", "content", "created_at"}]', 'App móvil'],
    ['POST', '/api/v1/chat', 'Chat con Fy', 'JWT', '100/min', '{"message", "conversation_id?"}', '{"response", "intent", "mood"}', 'App móvil -> Fy-Engine'],
]

ws1 = create_sheet(wb, 'API Gateway (8080)', '4472C4', headers, gateway_data, is_first=True)

# ============ FY-ENGINE (Puerto 8082) ============
engine_data = [
    ['GET', '/health', 'Health check', 'No', 'No', 'N/A', '{"status": "ok", "service": "fy-engine"}', 'API Gateway'],
    ['POST', '/chat', 'Chat principal con Fy', 'No', 'No', '{"user_id", "message", "context?"}', '{"response", "mood", "pii_detected", "intent", "analysis_performed"}', 'API Gateway'],
]

ws2 = create_sheet(wb, 'Fy-Engine (8082)', '70AD47', headers, engine_data)

# ============ FY-ANALYSIS (Puerto 9090) ============
analysis_data = [
    ['GET', '/health', 'Health check', 'No', '100/min', 'N/A', '{"status": "ok"}', 'Externo'],
    ['POST', '/analyze/url', 'Analizar URL (para Fy-Engine)', 'No', '100/min', '{"url"}', '{"risk_score", "verdict", "reasons", "brand_impersonation?"}', 'Fy-Engine'],
    ['POST', '/analyze/email', 'Analizar email (para Fy-Engine)', 'No', '100/min', '{"email"}', '{"risk_score", "verdict", "reasons"}', 'Fy-Engine'],
    ['POST', '/analyze/phone', 'Analizar teléfono (para Fy-Engine)', 'No', '100/min', '{"phone"}', '{"risk_score", "verdict", "reasons"}', 'Fy-Engine'],
    ['POST', '/api/v1/analyze', 'Análisis unificado (URL Engine)', 'No', '100/min', '{"url"}', '{"risk_score", "verdict", "checks", "metadata"}', 'Directo'],
    ['POST', '/api/v1/analyze/email', 'Analizar email (legacy)', 'No', '100/min', '{"email"}', '{"risk_score", "verdict"}', 'Directo'],
    ['POST', '/api/v1/analyze/url', 'Analizar URL (legacy)', 'No', '100/min', '{"url"}', '{"risk_score", "verdict"}', 'Directo'],
    ['POST', '/api/v1/analyze/phone', 'Analizar teléfono (legacy)', 'No', '100/min', '{"phone"}', '{"risk_score", "verdict"}', 'Directo'],
    ['POST', '/api/v1/analyze/batch', 'Análisis batch', 'No', '100/min', '{"urls", "emails", "phones"}', '[{"type", "value", "result"}]', 'Directo'],
    ['POST', '/api/v1/urlengine/check', 'URL Engine check', 'No', '100/min', '{"url"}', '{"risk_score", "sources", "metadata"}', 'Directo'],
    ['GET', '/api/v1/urlengine/status', 'Estado URL Engine', 'No', '100/min', 'N/A', '{"databases", "last_sync"}', 'Directo'],
    ['POST', '/api/v1/urlengine/sync', 'Sincronizar DBs', 'No', '100/min', 'N/A', '{"synced", "errors"}', 'Admin'],
]

ws3 = create_sheet(wb, 'Fy-Analysis (9090)', 'ED7D31', headers, analysis_data)

# ============ FLUJO DE COMUNICACIÓN ============
ws4 = wb.create_sheet('Flujo Comunicación')
flow_fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')

flow_headers = ['Paso', 'Origen', 'Destino', 'Endpoint', 'Descripción']
for col, header in enumerate(flow_headers, 1):
    cell = ws4.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = flow_fill
    cell.border = border

flow_data = [
    ['1', 'App Móvil', 'API Gateway', 'POST /auth/register', 'Usuario se registra'],
    ['2', 'App Móvil', 'API Gateway', 'POST /auth/send-code', 'Solicita código SMS'],
    ['3', 'App Móvil', 'API Gateway', 'POST /auth/verify', 'Verifica código -> recibe JWT'],
    ['4', 'App Móvil', 'API Gateway', 'POST /api/v1/chat', 'Envía mensaje a Fy'],
    ['5', 'API Gateway', 'Fy-Engine', 'POST /chat', 'Procesa mensaje con LLM'],
    ['6a', 'Fy-Engine', 'Fy-Analysis', 'POST /analyze/url', 'Si hay URL, analiza amenaza'],
    ['6b', 'Fy-Engine', 'Fy-Analysis', 'POST /analyze/email', 'Si hay email, analiza amenaza'],
    ['6c', 'Fy-Engine', 'Fy-Analysis', 'POST /analyze/phone', 'Si hay teléfono, analiza amenaza'],
    ['7', 'Fy-Analysis', 'Fy-Engine', 'Response', 'Devuelve risk_score y verdict'],
    ['8', 'Fy-Engine', 'API Gateway', 'Response', 'Devuelve respuesta de Fy'],
    ['9', 'API Gateway', 'App Móvil', 'Response', 'Muestra respuesta al usuario'],
]

for row_num, row_data in enumerate(flow_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_num, column=col_num, value=value)
        cell.border = border

# ============ RESUMEN PUERTOS ============
ws5 = wb.create_sheet('Puertos y DBs')
ports_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')

ports_headers = ['Servicio', 'Puerto Host', 'Puerto Interno', 'Base de Datos', 'Descripción']
for col, header in enumerate(ports_headers, 1):
    cell = ws5.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = ports_fill
    cell.border = border

ports_data = [
    ['API Gateway', '8080', '8080', 'PostgreSQL (trackfy_gateway)', 'Punto de entrada, auth, sesiones'],
    ['Fy-Engine', '8082', '8082', 'Redis (DB 0)', 'Orquestador LLM, guardrails'],
    ['Fy-Analysis', '9090', '9090', 'PostgreSQL (fy_threats)', 'Motor de análisis de amenazas'],
    ['PostgreSQL Gateway', '5433', '5432', 'trackfy_gateway', 'Usuarios, conversaciones'],
    ['PostgreSQL Threats', '5432', '5432', 'fy_threats', 'Whitelist, blacklist, amenazas'],
    ['Redis', '6379', '6379', 'DB 0: Fy-Engine, DB 1: Gateway', 'Cache, sesiones, rate limit'],
]

for row_num, row_data in enumerate(ports_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws5.cell(row=row_num, column=col_num, value=value)
        cell.border = border

# Ajustar anchos en todas las hojas
for ws in wb.worksheets:
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 40
    if len(list(ws.columns)) > 5:
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 45
        ws.column_dimensions['H'].width = 22

wb.save(r'c:\Users\Lenovo\Proyectos\GL\Trackfy\endpoints.xlsx')
print('Excel creado con todos los microservicios')
